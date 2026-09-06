"""Tabular exports. Synchronous today (capped by MAX_EXPORT_ROWS); the
`Export` value object is the seam where a queued job would hand back a
file later instead of streaming it now."""

import csv
import io
from dataclasses import dataclass
from typing import Any, Iterable, List, Sequence

from app.config import settings
from app.core.errors import ValidationError


@dataclass
class Export:
    filename: str
    content_type: str
    content: bytes


def _cell(value: Any) -> Any:
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def _rows(rows: Sequence[dict], columns: List[str]) -> List[List[Any]]:
    if len(rows) > settings.max_export_rows:
        raise ValidationError(f"Export exceeds the {settings.max_export_rows} row limit. Narrow the filters.", code="export_too_large")
    return [[_cell(r.get(c)) for c in columns] for r in rows]


def to_csv(name: str, columns: List[str], rows: Sequence[dict]) -> Export:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(columns)
    writer.writerows(_rows(rows, columns))
    return Export(f"{name}.csv", "text/csv; charset=utf-8", buf.getvalue().encode("utf-8"))


def to_xlsx(name: str, columns: List[str], rows: Sequence[dict], summary: dict | None = None) -> Export:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = name[:31]
    ws.append(columns)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in _rows(rows, columns):
        ws.append(row)
    for col in ws.columns:
        width = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(10, width + 2), 50)
    if summary:
        ws2 = wb.create_sheet("Summary")
        ws2.append(["metric", "value"])
        for k, v in summary.items():
            ws2.append([k, _cell(v)])
    out = io.BytesIO()
    wb.save(out)
    return Export(f"{name}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", out.getvalue())


def to_pdf(name: str, title: str, columns: List[str], rows: Sequence[dict], summary: dict | None = None, filters: str = "") -> Export:
    from fpdf import FPDF

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 8, title, new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 6, f"Filters: {filters}", new_x="LMARGIN", new_y="NEXT")
    if summary:
        pdf.ln(2)
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(0, 6, "Summary", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 9)
        for k, v in summary.items():
            pdf.cell(0, 5, f"{k.replace('_', ' ')}: {_cell(v)}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)
    usable = pdf.w - 2 * pdf.l_margin
    col_w = usable / max(1, len(columns))
    pdf.set_font("Helvetica", "B", 9)
    for c in columns:
        pdf.cell(col_w, 7, str(c)[:22], border=1)
    pdf.ln()
    pdf.set_font("Helvetica", "", 8)
    for row in _rows(rows, columns):
        for value in row:
            pdf.cell(col_w, 6, str(value)[:26], border=1)
        pdf.ln()
    return Export(f"{name}.pdf", "application/pdf", bytes(pdf.output()))


def export(fmt: str, name: str, title: str, columns: List[str], rows: Sequence[dict], summary: dict | None = None, filters: str = "") -> Export:
    if fmt == "csv":
        return to_csv(name, columns, rows)
    if fmt == "xlsx":
        return to_xlsx(name, columns, rows, summary)
    if fmt == "pdf":
        return to_pdf(name, title, columns, rows, summary, filters)
    raise ValidationError("format must be csv, xlsx or pdf")
